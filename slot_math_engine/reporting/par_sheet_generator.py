import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class PARSheetGenerator:
    def __init__(self, calculator, reel_strip):
        self.calc = calculator
        self.reel = reel_strip
        
    def generate_full_par_sheet(self, output_path: str):
        wb = Workbook()
        
        ws_summary = wb.active
        ws_summary.title = "Game Summary"
        self._populate_summary(ws_summary)
        
        ws_reels = wb.create_sheet("Reel Strips")
        self._populate_reel_strips(ws_reels)
        
        ws_probs = wb.create_sheet("Symbol Probabilities")
        self._populate_symbol_probs(ws_probs)
        
        ws_paytable = wb.create_sheet("Pay Table Math")
        self._populate_paytable(ws_paytable)
        
        ws_hits = wb.create_sheet("Hit Frequency")
        self._populate_hit_frequency(ws_hits)
        
        ws_vol = wb.create_sheet("Volatility")
        self._populate_volatility(ws_vol)
        
        wb.save(output_path)
    
    def _populate_summary(self, ws):
        headers = ["Parameter", "Value", "Notes"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        data = [
            ["Game Name", "Mathematician Demo Slot", "5-Reel Video Slot"],
            ["Reel Configuration", "5 x 3 (5 Reels, 3 Rows)", "Standard Layout"],
            ["Total Virtual Stops", f"{self.reel.reel_stops}", "Per reel"],
            ["Total Combinations", f"{np.prod(self.reel.reel_stops):,}", "Exact mathematical space"],
            ["Paylines", "20", "Fixed paylines"],
            ["Bet per Spin", "20 credits", "1 credit per line"],
            ["Theoretical RTP", f"{self.calc.rtp if hasattr(self.calc, 'rtp') else 0:.4f}%", "Calculated"],
            ["Hit Frequency", f"{self.calc.hit_frequency if hasattr(self.calc, 'hit_frequency') else 0:.4f}%", "Winning spins percentage"],
            ["Volatility Index", f"{self.calc.std_dev if hasattr(self.calc, 'std_dev') else 0:.4f}", "Standard deviation of returns"]
        ]
        
        for row in data:
            ws.append(row)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def _populate_symbol_probs(self, ws):
        headers = ["Symbol", "Reel 1", "Reel 2", "Reel 3", "Reel 4", "Reel 5", 
                   "5oak Prob", "4oak Prob", "3oak Prob", "EV Contribution"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for sym_name, sym_data in self.reel.symbols.items():
            row = [sym_name]
            
            reel_probs = []
            for reel_idx in range(5):
                prob = self.reel.get_symbol_probability(reel_idx, sym_name)
                reel_probs.append(prob)
                row.append(f"{prob:.6f}")
            
            prob_5oak = np.prod(reel_probs)
            
            prob_4oak = 0
            for miss_reel in range(5):
                prob = 1
                for reel_idx in range(5):
                    if reel_idx == miss_reel:
                        prob *= (1 - reel_probs[reel_idx])
                    else:
                        prob *= reel_probs[reel_idx]
                prob_4oak += prob
            
            prob_3oak = 0
            from itertools import combinations
            for match_reels in combinations(range(5), 3):
                prob = 1
                for reel_idx in range(5):
                    if reel_idx in match_reels:
                        prob *= reel_probs[reel_idx]
                    else:
                        prob *= (1 - reel_probs[reel_idx])
                prob_3oak += prob
            
            row.extend([
                f"{prob_5oak:.8e}",
                f"{prob_4oak:.8e}",
                f"{prob_3oak:.8e}"
            ])
            
            payouts = sym_data['payout']
            payout_5 = payouts[5] if len(payouts) > 5 else 0
            payout_4 = payouts[4] if len(payouts) > 4 else 0
            payout_3 = payouts[3] if len(payouts) > 3 else 0
            
            ev_contrib = (prob_5oak * payout_5 + 
                         prob_4oak * payout_4 + 
                         prob_3oak * payout_3) / 20
            
            row.append(f"{ev_contrib:.6f}")
            ws.append(row)
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
    
    def _populate_reel_strips(self, ws):
        ws.append(["REEL STRIP CONFIGURATION"])
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        for reel_idx in range(5):
            start_row = ws.max_row + 2
            ws.append([f"REEL {reel_idx + 1}"])
            ws.merge_cells(f'A{start_row}:C{start_row}')
            ws[f'A{start_row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'A{start_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            ws.append(["Stop Position", "Symbol", "Weight", "Cumulative", "Probability"])
            
            reel_data = self.reel.cumulative_weights[reel_idx]
            cumsum = 0
            
            for i, (sym, weight) in enumerate(zip(reel_data['symbols'], reel_data['weights'])):
                cumsum += weight
                prob = weight / reel_data['total']
                ws.append([i, sym, weight, cumsum, f"{prob:.4%}"])
    
    def _populate_paytable(self, ws):
        headers = ["Symbol", "3 of a Kind", "4 of a Kind", "5 of a Kind",
                   "Prob 3oak", "Prob 4oak", "Prob 5oak", 
                   "EV 3oak", "EV 4oak", "EV 5oak", "Total EV"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        total_ev = 0
        
        for sym_name, sym_data in self.reel.symbols.items():
            if not sym_data.get('payout') or len(sym_data['payout']) <= 3:
                continue

            p3 = sym_data['payout'][3] if len(sym_data['payout']) > 3 else 0
            p4 = sym_data['payout'][4] if len(sym_data['payout']) > 4 else 0
            p5 = sym_data['payout'][5] if len(sym_data['payout']) > 5 else 0

            if p3 == 0 and p4 == 0 and p5 == 0:
                continue
            
            reel_probs = [self.reel.get_symbol_probability(i, sym_name) for i in range(5)]
            
            is_scatter = sym_name == "SCATTER"
            
            if not is_scatter:
                prob_5 = np.prod(reel_probs)
                prob_4 = np.prod(reel_probs[:4]) * (1 - reel_probs[4])
                prob_3 = np.prod(reel_probs[:3]) * (1 - reel_probs[3])
            else:
                prob_5 = np.prod(reel_probs)
                prob_4 = np.prod(reel_probs[:4]) * (1 - reel_probs[4])
                prob_3 = np.prod(reel_probs[:3]) * (1 - reel_probs[3])

            ev_3 = prob_3 * p3 / 20
            ev_4 = prob_4 * p4 / 20
            ev_5 = prob_5 * p5 / 20
            ev_total = ev_3 + ev_4 + ev_5
            total_ev += ev_total
            
            row = [
                sym_name,
                p3, p4, p5,
                f"{prob_3:.6e}",
                f"{prob_4:.6e}",
                f"{prob_5:.6e}",
                f"{ev_3:.6f}",
                f"{ev_4:.6f}",
                f"{ev_5:.6f}",
                f"{ev_total:.6f}"
            ]
            ws.append(row)
        
        ws.append([])
        ws.append(["TOTAL THEORETICAL RTP", f"{total_ev*100:.4f}%", "", "", "", "", "", "", "", "", ""])
        ws[f'A{ws.max_row}'].font = Font(bold=True)
        ws[f'B{ws.max_row}'].font = Font(bold=True, color="FF0000")
    
    def _populate_hit_frequency(self, ws):
        ws.append(["HIT FREQUENCY ANALYSIS"])
        ws['A1'].font = Font(bold=True, size=14)
        
        categories = [
            ("Loss", 0, 0),
            ("Push (0-1x)", 1, 20),
            ("Small Win (1-5x)", 21, 100),
            ("Medium Win (5-20x)", 101, 400),
            ("Big Win (20-100x)", 401, 2000),
            ("Mega Win (100x+)", 2001, 100000)
        ]
        
        headers = ["Category", "Min Credits", "Max Credits", "Hit Probability", "Contribution to RTP"]
        ws.append(headers)
        
        for cat_name, min_cred, max_cred in categories:
            ws.append([cat_name, min_cred, max_cred, "Calculated from sim", "Calculated from sim"])
    
    def _populate_volatility(self, ws):
        ws.append(["VOLATILITY ANALYSIS"])
        
        variance = getattr(self.calc, 'variance', 0)
        std_dev = getattr(self.calc, 'std_dev', 0)
        hit_freq = getattr(self.calc, 'hit_frequency', 0)
        rtp = getattr(self.calc, 'rtp', 0)
        
        metrics = [
            ["Standard Deviation", f"{std_dev:.4f}"],
            ["Variance", f"{variance:.4f}"],
            ["Coefficient of Variation", f"{std_dev/rtp if rtp else 0:.4f}"],
            ["Hit Frequency (%)", f"{hit_freq:.4f}"],
            ["Expected Spins Between Wins", f"{100/hit_freq if hit_freq else 0:.2f}"],
            ["Max Win (5 Wilds)", "2000x"],
            ["Risk of Ruin (100 unit bankroll)", "Calculated"]
        ]
        
        for metric, value in metrics:
            ws.append([metric, value])